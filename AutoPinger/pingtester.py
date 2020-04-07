import pyspeedtest
import openpyxl
from pathlib import Path
import os
import time
import datetime


servers = ['www.google.com','www.twitch.tv','www.github.com','www.egov.kz','www.speedtest.com']

wbname = Path(os.getcwd())
wbname= os.path.join(wbname,Path("pingdata.xlsx"))
wb = openpyxl.load_workbook(wbname)

delay = 15*60
index = 2

def Measure():
    for i in range(len(servers)):
        print('Testing connection with:',servers[i])
        ws = wb.worksheets[i]
        st = pyspeedtest.SpeedTest(servers[i])
        st.chooseserver(servers[i])
        dnl = float(st.download())
        ping = st.ping()
        dnl /= 1024
        c = ws.cell(index,1,datetime.datetime.now())
        c = ws.cell(index,3,'{:.2f}'.format(round(dnl,2))+' kb/s')
        c = ws.cell(index,2,'{:.1f}'.format(round(ping,1))+' mls')

while True:
    print('Time:',datetime.datetime.now())
    Measure() 
    wb.save(wbname)
    index+=1
    print('Everything is tested, going back to sleep')
    time.sleep(delay)
    
